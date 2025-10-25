import pandas as pd
import pickle
import openpyxl
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix



def model(name):
    
    file = pd.ExcelFile(name)

    dict={}
    # getting the sheetnames
    sheet_names = file.sheet_names
    for s in sheet_names:
        dict[s]=s.replace("-","_") if "-" in s else s
    # print(dict)

    new=[]
    for old in dict.values():
        new.append(old)

    ind=0
    for sheet_name in sheet_names:
        sheet = file.parse(sheet_name)
        date=sheet.columns[0]
        sheet.columns=sheet.iloc[0]
        globals()[new[ind]] = sheet
        ind+=1

    global Fee_Earnings
    # Filter out rows with non-numeric values in the "Returns" column
    valid_rows = pd.to_numeric(Fee_Earnings['Returns'], errors='coerce').notna()
    Fee_Earnings = Fee_Earnings[valid_rows]

    # Convert the "Returns" column to integers
    Fee_Earnings['Returns'] = Fee_Earnings['Returns'].astype(int)


    product_name=Fee_Earnings["Name"]

    Product_return=Fee_Earnings["Returns"]

    Product_return=Product_return.astype(int)


    X_train, X_test, Y_train, Y_test = train_test_split(
        product_name, Product_return, test_size=0.2, random_state=3)

    feature_extraction = TfidfVectorizer(
        min_df=1, stop_words='english', lowercase=True)

    X_train_features = feature_extraction.fit_transform(X_train)
    X_test_features = feature_extraction.transform(X_test)

    model = LogisticRegression()

    # Training the Logistic Regression model with the training data
    model.fit(X_train_features, Y_train)

    import pickle as p
    with open('cmodel_pkl', 'wb') as files:
        p.dump(model , files)

    with open("feature_extraction", 'wb') as files:
        p.dump(feature_extraction, files)

    # Prediction on training data
    prediction_on_training_data = model.predict(X_train_features)
    accuracy_on_training_data = accuracy_score(Y_train, prediction_on_training_data)
    precision_on_training_data = precision_score(Y_train, prediction_on_training_data)
    recall_on_training_data = recall_score(Y_train, prediction_on_training_data)
    f1_on_training_data = f1_score(Y_train, prediction_on_training_data)
    confusion_matrix_train = confusion_matrix(Y_train, prediction_on_training_data)

    print('Training Set Metrics:')
    print('Accuracy:', accuracy_on_training_data)
    print('Precision:', precision_on_training_data)
    print('Recall:', recall_on_training_data)
    print('F1 Score:', f1_on_training_data)
    print('Confusion Matrix:')
    print(confusion_matrix_train)

    # Prediction on test data
    prediction_on_test_data = model.predict(X_test_features)
    accuracy_on_test_data = accuracy_score(Y_test, prediction_on_test_data)
    precision_on_test_data = precision_score(Y_test, prediction_on_test_data)
    recall_on_test_data = recall_score(Y_test, prediction_on_test_data)
    f1_on_test_data = f1_score(Y_test, prediction_on_test_data)
    confusion_matrix_test = confusion_matrix(Y_test, prediction_on_test_data)

    print('\nTest Set Metrics:')
    print('Accuracy:', accuracy_on_test_data)
    print('Precision:', precision_on_test_data)
    print('Recall:', recall_on_test_data)
    print('F1 Score:', f1_on_test_data)
    print('Confusion Matrix:')
    print(confusion_matrix_test)

    return True




def prediction(name):
    # name="1698492451662-Fee-Earnings-85f7091e-f0a8-4a18-b86c-82f1b689cc09-XLSX.xlsx"
    with open('cmodel_pkl', 'rb') as f:
        lr = pickle.load(f)

    with open('feature_extraction','rb') as f:
        feature_extraction=pickle.load(f)

    products="product_details.xlsx"

    sht=pd.read_excel(products,sheet_name="Product_details")
    file = pd.ExcelFile(name)

    dict={}
    # getting the sheetnames
    sheet_names = file.sheet_names
    for s in sheet_names:
        dict[s]=s.replace("-","_") if "-" in s else s
    # print(dict)

    new=[]
    for old in dict.values():
        new.append(old)

    ind=0
    for sheet_name in sheet_names:
        sheet = file.parse(sheet_name)
        date=sheet.columns[0]
        sheet.columns=sheet.iloc[0]
        globals()[new[ind]] = sheet
        ind+=1

    product_name=sht["Product_Name"]
    # print(product_name)

    X_train_features = feature_extraction.transform(product_name)

    pred=lr.predict(X_train_features)

    sht['result'] = pred
    if not pd.ExcelFile("product_details.xlsx").sheet_names.__contains__("Results"):
        with pd.ExcelWriter(products, engine='openpyxl', mode='a') as writer:
            sht.to_excel(writer, sheet_name='Results', index=False)
    else:
        # Delete the Results sheet
        wb = openpyxl.load_workbook(products)
        ws = wb.get_sheet_by_name("Results")
        wb.remove_sheet(ws)
        wb.save(products)
        with pd.ExcelWriter(products, engine='openpyxl', mode='a') as writer:
            sht.to_excel(writer, sheet_name='Results', index=False)
            
    return True

# prediction(name="1698492451662-Fee-Earnings-85f7091e-f0a8-4a18-b86c-82f1b689cc09-XLSX.xlsx")

